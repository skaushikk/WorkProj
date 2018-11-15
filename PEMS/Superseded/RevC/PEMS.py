import sys
import os
#import argeparse
import OrcFxAPI
from openpyxl import load_workbook
from openpyxl import Workbook
import csv

def PEMS(FPath, FName, TTen, TBM, EndA, WinchB, Content):
    basefilename = FPath + "\\" + 'PEMS_Base.dat'
    model = OrcFxAPI.Model(basefilename)
    
    # Grab an individual Winch and pipeline
    pipelineWinch = model['Winch1']
    pipeline = model['Pipeline']
    
    pipeline.EndAX = EndA[0]
    pipeline.EndAZ = EndA[2]
    pipeline.EndADeclination = EndA[3]
   # Grab X position of your Winch
    
    pipelineWinch.ConnectionX[0]=WinchB[0]
    pipelineWinch.ConnectionZ[0]=WinchB[1]
    
    pipelineWinchX = pipelineWinch.ConnectionX[0]
    pipelineWinchTension = pipelineWinch.StageValue[0]
        
    if Content == 'Flooded':
        pipeline.ContentsDensity = 1.025
    if Content == 'Empty':
        pipeline.contentsDensity = 0
    
    # Calculate statics
    model.CalculateStatics()
    
    # extract tension and bending moment results
    pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
    pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
    print("Initial Tension= ",round(pipeline_tension,1),"Initial BM= ",abs(round(pipeline_bending_moment,2)))
    print("Initial WinchX=  ",pipelineWinchX)
    # define target values
    target_tension = TTen
    target_bending_moment = TBM
    
    # calculate initial error
    tension_error = abs(target_tension - pipeline_tension)
    bending_moment_error = abs(target_bending_moment + pipeline_bending_moment)
    print(tension_error)
   
   # logic loop
    loop_count = 0
    i=2
    XStep=2.5
    BMloop=[0,0]
    while tension_error > 0.1 or bending_moment_error > 0.01:
            TensionDiff = target_tension-pipeline_tension
            pipelineWinchTension = pipelineWinch.StageValue[0]
            pipelineWinch.StageValue[0] = pipelineWinchTension + TensionDiff
            model.CalculateStatics()
    
            pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
            pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
            # re-calculate errors
            tension_error = abs(target_tension + pipeline_tension)
            bending_moment_error = abs(target_bending_moment + pipeline_bending_moment)
    
            #Check the moment to find the X-step direction
            
            BMDiff = target_bending_moment + pipeline_bending_moment
            pipelineWinchX -= BMDiff/abs(BMDiff)*XStep
    
    
            # assign the Winch with new end positions
            pipelineWinch.ConnectionX[0] = pipelineWinchX
    
            # re-calculate statics, extract results
            model.CalculateStatics()
            pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
            pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
            # re-calculate errors
            tension_error = abs(target_tension - pipeline_tension)
            bending_moment_error = abs(target_bending_moment + pipeline_bending_moment)
            
            loop_count += 1
            if loop_count > 5000:
                break
            print("loop number ",loop_count,"  Ten= ",round(pipeline_tension,5),"  BM= ",round(pipeline_bending_moment,5),"WinchX=",pipelineWinchX)
            
            BMloop.append(abs(pipeline_bending_moment))
            if round(BMloop[i],3) == round(BMloop[i-2],3):
                XStep = XStep/10
    
            i=i+1
    
    # Save .dat file once the 'while' loop completes
    model.SaveData(FName)
    return 0;

def main():
        
    filePath = sys.argv[1]
    wbName = sys.argv[2]
    
    wbName = filePath + "\\" + wbName
    wb = load_workbook(wbName,data_only=True)
    ws = wb.get_sheet_by_name('Pre-processing')  
    wr = wb.get_sheet_by_name('Report')
    EndA = [ws['A33'].value,ws['B33'].value,ws['C33'].value,ws['D29'].value+90]
    WinchB = [ws['E33'].value,ws['G33'].value]
    content = ws['D27'].value
    LCFolder = ws['D28'].value
    PipeID = ws['D25'].value
    CSVName = str(filePath) + "\\" + str(PipeID) + "_" + LCFolder + ".csv"
    directory = filePath+ "\\"+LCFolder    
   
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    i=1
    while i<6:       
        FNameID = ws.cell(row=38+i,column=1).value
        FName1= str(filePath) + "\\" + str(LCFolder) + "\\"+ str(FNameID)
        TargetTen=ws.cell(row=38+i,column=3).value
        TargetBM=ws.cell(row=38+i,column=4).value
        PEMS(filePath, FName1, TargetTen, TargetBM, EndA, WinchB, content)
        i=i+1
    
#    with open(CSVName, 'w',newline='') as csvfile:
#        c = csv.writer(csvfile,quoting = csv.QUOTE_MINIMAL)
#        for r in wr.rows:
#            c.writerow([str(cell.value) for cell in r])   
#            
#    with open(CSVName, 'w',newline='') as csvfile:
#        c = csv.writer(csvfile)
#        c.writerows(wr.row_values(row) for row in range(wr.nrows))  
            
            
if __name__ == "__main__":
    main ()
 
