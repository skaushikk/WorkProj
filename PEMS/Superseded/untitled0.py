import sys
import os
#import argeparse
import OrcFxAPI
from openpyxl import load_workbook
from openpyxl import Workbook
import csv

def PEMS(FileData, Tloads, PipeEndA, WinchB, Content, PIP_l):
    
    if Content == 'Flooded':
        PEMS_Base_File = 'PEMS_base_F.dat'
    if Content == 'Empty':
        PEMS_Base_File = 'PEMS_base_E.dat'
    
    basefilename = FileData[0] + "\\" + PEMS_Base_File
    model = OrcFxAPI.Model(basefilename)
    
    ## Define pipeline and winch end coordinates
    
    # Grab an individual Winch and pipeline
    pipelineWinch = model['Winch1']
    pipeline = model['Pipeline']
    
    pipeline.EndAX = PipeEndA[0]
    pipeline.EndAZ = PipeEndA[2]
    pipeline.EndADeclination = PipeEndA[3]
   
   # Grab X position of your Winch
    
    pipelineWinch.ConnectionX[0]=WinchB[0]
    pipelineWinch.ConnectionZ[0]=WinchB[1]
    
    pipelineWinchX = pipelineWinch.ConnectionX[0]
    pipelineWinchTension = pipelineWinch.StageValue[0]
        
    ## pipeline content definition
    
#    if PIP_l == 'NO':
#        if Content == 'Flooded':
#            pipeline.ContentsDensity = 1.025
#        if Content == 'Empty':
#            pipeline.contentsDensity = 0
#    if PIP_l == 'YES':

	
    
    # Calculate statics
    model.CalculateStatics()
    
    # extract tension and bending moment results
    pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
    pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
    print("Initial Tension= ",round(pipeline_tension,1),"Initial BM= ",abs(round(pipeline_bending_moment,2)))
    print("Initial WinchX=  ",pipelineWinchX)
    # define target values
    target_tension = Tloads[0]
    target_bending_moment = Tloads[1]
    
    # calculate initial error
    tension_error = abs(target_tension - pipeline_tension)
    bending_moment_error = abs(target_bending_moment + pipeline_bending_moment)
    print(tension_error)
   
   # logic loop
    loop_count = 0
    i=2
    XStep=2.5
    BMloop=[0,0]
    while tension_error > 0.1 or bending_moment_error > 0.1:
            TensionDiff = target_tension-pipeline_tension
            pipelineWinchTension = pipelineWinch.StageValue[0]
            pipelineWinch.StageValue[0] = pipelineWinchTension + TensionDiff
            model.CalculateStatics()
    
            pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
            pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
            # re-calculate errors
            tension_error = abs(target_tension + pipeline_tension)
            bending_moment_error = abs(target_bending_moment + pipeline_bending_moment)
    
            #Check the moment to find the X-step direction
            
            BMDiff = target_bending_moment + pipeline_bending_moment
            pipelineWinchX -= BMDiff/abs(BMDiff)*XStep
    
    
            # assign the Winch with new end positions
            pipelineWinch.ConnectionX[0] = pipelineWinchX
    
            # re-calculate statics, extract results
            model.CalculateStatics()
            pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
            pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
            # re-calculate errors
            tension_error = abs(target_tension - pipeline_tension)
            bending_moment_error = abs(target_bending_moment + pipeline_bending_moment)
            
            loop_count += 1
            if loop_count > 5000:
                break
            print("loop number ",loop_count,"  Ten= ",round(pipeline_tension,5),"  BM= ",round(pipeline_bending_moment,5),"WinchX=",pipelineWinchX)
            
            BMloop.append(abs(pipeline_bending_moment))
            if round(BMloop[i],2) == round(BMloop[i-2],2):
                XStep = XStep/10
    
            i=i+1
    
    # Save .dat file once the 'while' loop completes
    model.SaveData(FileData[1])
    return 0;

def main():
        
    filePath = sys.argv[1]
    wbName = sys.argv[2]
    
    wbName = filePath + "\\" + wbName
    wb = load_workbook(wbName,data_only=True)
    ws = wb.get_sheet_by_name('Pre-processing')  
    wr = wb.get_sheet_by_name('Report')
    PipeEndA = [ws['A33'].value,ws['B33'].value,ws['C33'].value,ws['D29'].value+90]
    WinchB = [ws['E33'].value,ws['G33'].value]
    content = ws['D27'].value
    LCFolder = str(ws['D29'].value) + "_" + str(ws['D26'].value) + "_" + str(ws['D27'].value)
    PipeID = ws['D25'].value
    CSVName = str(filePath) + "\\" + str(PipeID) + "_" + LCFolder + ".csv"
    directory = filePath+ "\\"+LCFolder
    PIP_l = ws['F25'].value
    if not os.path.exists(directory):
        os.makedirs(directory)


    i=1
    
    while i<6:       
        FNameID = ws.cell(row=38+i,column=1).value
        FName1= str(filePath) + "\\" + str(LCFolder) + "\\"+ str(FNameID)
        FileData = [filePath, FName1]
        
        TargetTen=ws.cell(row=38+i,column=3).value
        TargetBM=ws.cell(row=38+i,column=4).value
        Tloads = [TargetTen, TargetBM]
        PEMS(FileData, Tloads, PipeEndA, WinchB, content, PIP_l)
        i=i+1
#    
    with open(CSVName, 'w',newline='') as csvfile:
        c = csv.writer(csvfile,quoting = csv.QUOTE_MINIMAL)
        for r in wr.rows:
            c.writerow([str(cell.value) for cell in r])
            
#            
#    with open(CSVName, 'w',newline='') as csvfile:
#        c = csv.writer(csvfile)
#        c.writerows(wr.row_values(row) for row in range(wr.nrows))  
       
#    TempFile = filePath + "\\" + 'TEMP.bat'
#    os.remove(TempFile)       
           
if __name__ == "__main__":
    main ()
    
