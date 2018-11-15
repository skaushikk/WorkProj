import sys
import os
from openpyxl import load_workbook
import csv


def main():
        
    filePath = sys.argv[1]
    wbName = sys.argv[2]
    
    wbName = filePath + "\\" + wbName
    wb = load_workbook(wbName,data_only=True)
    ws = wb.get_sheet_by_name('Pre-processing')  
    wr = wb.get_sheet_by_name('Report')
    PipeEndA = [ws['A33'].value,ws['B33'].value,ws['C33'].value,ws['D29'].value+90]
    WinchB = [ws['E33'].value,ws['G33'].value]
    content = ws['D27'].value
    LCFolder = str(ws['D29'].value) + "_" + str(ws['D26'].value) + "_" + str(ws['D27'].value)
    PipeID = ws['D25'].value
    CSVName = str(filePath) + "\\" + str(PipeID) + "_" + LCFolder + ".csv"
    directory = filePath+ "\\"+LCFolder
    PIP_l = ws['F25'].value
    if not os.path.exists(directory):
        os.makedirs(directory)


#    with open(CSVName, 'w',newline='') as csvfile:
#        c = csv.writer(csvfile,quoting=csv.QUOTE_MINIMAL)
#        for r in wr.rows:
#            c.writerow([str(cell.value) for cell in r])
            
#            
    with open(CSVName, 'w',newline='') as csvfile:
        c = csv.writer(csvfile)
        c.writerows(wr.row_values(row) for row in range(wr.nrows))  
       
#    TempFile = filePath + "\\" + 'TEMP.bat'
#    os.remove(TempFile)       
           
if __name__ == "__main__":
    main ()
    
