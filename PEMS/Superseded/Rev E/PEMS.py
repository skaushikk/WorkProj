# -*- coding: utf-8 -*-
"""
Created on Fri Sep  2 14:36:10 2016

@author: ksirvole
"""
import sys
import os
import OrcFxAPI
from openpyxl import load_workbook
import warnings
import math

def PEMS(Fileinfo, Tloads, PipeEndA, WinchB, Contents, Tol):
    
    # Set the input file based on contents
    if Contents == 'Flooded':
        PEMS_Base_File = 'PEMS_base_F.dat'
    if Contents == 'Empty':
        PEMS_Base_File = 'PEMS_base_E.dat'
    
    basefilename = Fileinfo[0] + "\\" + PEMS_Base_File
                           
    if not os.path.exists(basefilename):
        sys.exit(1)
        
#    directory = filePath+ "\\"+LCFolder
#    if not os.path.exists(directory):
#        os.makedirs(directory)
    
    model = OrcFxAPI.Model(basefilename)
    
    
    ## Define pipeline and winch end coordinates
    
    # Grab an individual Winch and pipeline
    pipelineWinch = model['Winch1']
    pipeline = model['Pipeline']
    
   # Assign pipe endA corrdinates
    pipeline.EndAX = PipeEndA[0]
    pipeline.EndAZ = PipeEndA[2]
    pipeline.EndADeclination = PipeEndA[3]
   
   # assign initial end coordinates for Winch
    
    pipelineWinch.ConnectionX[0]=WinchB[0]
    pipelineWinch.ConnectionZ[0]=WinchB[1]
    
    pipelineWinchX = pipelineWinch.ConnectionX[0]
    pipelineWinchTension = pipelineWinch.StageValue[0]
        
   
    # Calculate statics
    model.CalculateStatics()
    
    # extract tension and bending moment results
    pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
    pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
    print("\n")
    print("Initial Tension= ",round(pipeline_tension,1),"Initial BM= ",abs(round(pipeline_bending_moment,2)),"Initial WinchX=  ",round(pipelineWinchX,3))
   
    # define target values
    target_tension = Tloads[0]
    target_bending_moment = Tloads[1]

    print("Targer Tension= ",round(target_tension,1),"Target BM= ",abs(round(target_bending_moment,2)),"\n")
    
   # calculate initial error
    TensionError = abs(target_tension - pipeline_tension)
    BMError = abs(target_bending_moment + pipeline_bending_moment)
   # print(TensionError)
   
   # logic loop
    loop_count = 0
    i=2
    XStep=2
    precision = int(-math.log10(float(Tol)))
    BMloop=[0,0]
    while TensionError > 0.1 or BMError > 1/10**(precision-1):
            TensionDiff = target_tension-pipeline_tension
            pipelineWinchTension = pipelineWinch.StageValue[0]
            pipelineWinch.StageValue[0] = pipelineWinchTension + TensionDiff
            model.CalculateStatics()
    
            pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
            pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
    # re-calculate errors
            TensionError = abs(target_tension + pipeline_tension)
            BMError = abs(target_bending_moment + pipeline_bending_moment)
    
    #Check the moment to find the X-step direction
            
            BMDiff = target_bending_moment + pipeline_bending_moment
            pipelineWinchX -= BMDiff/abs(BMDiff)*XStep
    
    
    # assign the Winch with new end positions
            pipelineWinch.ConnectionX[0] = pipelineWinchX
    
    # re-calculate statics, extract results
            model.CalculateStatics()
            pipeline_tension = pipeline.StaticResult('Effective Tension', OrcFxAPI.oeEndA)
            pipeline_bending_moment = pipeline.StaticResult('y-Bend Moment', OrcFxAPI.oeEndA)
    
    # re-calculate errors
            TensionError = abs(target_tension - pipeline_tension)
            BMError = abs(target_bending_moment + pipeline_bending_moment)
            
            loop_count += 1
            if loop_count > 200:
                break
            print("Iteration ",loop_count,"  Ten= ",round(pipeline_tension,3),"  BM= ",round(pipeline_bending_moment,3),"WinchX=",round(pipelineWinchX,3))
            
            BMloop.append(abs(pipeline_bending_moment))
            
     # Reduce step size as it gets closer to the target values of WinchX       
            if round(BMloop[i],precision) == round(BMloop[i-2],precision):
                XStep = XStep/10
    
            i=i+1
    
    # Save .dat file once the 'while' loop completes
    model.SaveData(Fileinfo[1])
    return 0;

    # define main body of code to extract data from spreadsheet
def main():
        
    try:
        filePath = sys.argv[1]
        wbName = sys.argv[2]
    except IndexError:
        print("Spreadsheet NOT found. Using default.")        
        filePath = 'Q:/Subsea/Design Engineering/Inst Analysis/10. Consolidation/PEMS/TEST'
        wbName = "PEMS.xlsm"
    else:
        print ("Spreadsheet data found.")
    
    warnings.simplefilter("ignore")
    
    wbName = filePath + "\\" + wbName
    wb = load_workbook(wbName,data_only=True)
    ws = wb.get_sheet_by_name('Pre-processing')  
    wr = wb.get_sheet_by_name('Report')
    PipeEndA = [ws['F36'].value, ws['G36'].value, ws['H36'].value, ws['F26'].value+90]
    WinchB = [ws['F39'].value,ws['H39'].value]
    PipeID = ws['F23'].value
    Hangoff = ws['F24'].value
    contents = ws['F25'].value
    Hangoffangle = ws['F26'].value
    LCFolder = str(Hangoffangle) + "_" + str(Hangoff) + "_" + str(contents)
    Tol = ws['F31'].value
    
    # CSVName = str(filePath) + "\\" + str(PipeID) + "_" + LCFolder + ".csv"
    
    directory = filePath+ "\\"+LCFolder
    if not os.path.exists(directory):
        os.makedirs(directory)

    i=0
    rstart = 43
    while i<5:       
        FNameID = ws.cell(row=rstart+i,column=2).value
        FName1= str(filePath) + "\\" + str(LCFolder) + "\\"+ str(FNameID)
        FileinfoData = [filePath, FName1]
        
        TargetTen=ws.cell(row=rstart+i,column=4).value
        TargetBM=ws.cell(row=rstart+i,column=5).value
        Tloads = [TargetTen, TargetBM]
        PEMS(FileinfoData, Tloads, PipeEndA, WinchB, contents,Tol)
        i=i+1

           
if __name__ == "__main__":
    main ()
    
